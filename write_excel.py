import pyodbc
import pathlib
import time
import teradatasql
import jaydebeapi as jdbc_driver
from socket import gethostname
from datetime import datetime
import pandas as pd
import os
from numpy.ma.core import append
from tabulate import tabulate
from simple_colors import *
import openpyxl
import numpy as np
from robot.libraries.BuiltIn import BuiltIn
import configparser
from openpyxl.styles import PatternFill
from openpyxl.styles import Font, Color
# from openpyxl.styles.colors import Font, Color
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
import os
import json
import sys

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__))).replace('\\', '/')
sys.path.append(BASE_DIR)


def write_excel_file(file_path, sheet_name, df, mismatch_score_cell_id, missmatch_l2=[[],[]]):
    # Combine base directory with the file path
    file_path = documents_path + file_path
    # Extract the directory path from the full file path
    directory = os.path.dirname(file_path)
    # Create the directory if it doesn't exist
    if not os.path.exists(directory):
        os.makedirs(directory)
    # Check if the file exists and load it, otherwise create a new workbook
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
    else:
        wb = Workbook()
    # If 'Sheet' exists and is empty, delete it
    if 'Sheet' in wb.sheetnames:
        delete_empty_sheet(wb, 'Sheet')
    # Check if the sheet already exists
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]  # If the sheet exists, just use the existing one
    else:
        # If the sheet doesn't exist, create a new one
        sheet = wb.create_sheet(sheet_name)
        # Append Header
        sheet.append(list(df))
        # Populating the Excel Sheet with Data
        for row in df.iterrows():
            sheet.append(row[1].tolist())

    # Set column widths for readability
    ws = sheet
    dim_holder = DimensionHolder(worksheet=ws)
    for col in range(ws.min_column, ws.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=20)
    ws.column_dimensions = dim_holder
    # Adding mismatch data if provided

    if len(missmatch_l2[0]) != 0:
        sheet.append([''])
        sheet.append(['MISSING L2 CATEGORIES'])
        for i in missmatch_l2[0]:
            sheet.append([i])

    if len(missmatch_l2[1]) != 0:
        sheet.append([''])
        sheet.append(['NEWLY ADDED L2 CATEGORIES'])
        for i in missmatch_l2[1]:
            sheet.append([i])
    # Highlight header with yellow background
    row, col = df.shape
    cell_ids = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U',
                'V', 'W', 'X', 'Y', 'Z']
    # Check link for color code:  https://openpyxl.readthedocs.io/en/stable/styles.html
    temp_font = Font(color="FFFFFF")  # White font color
    header_colour_code = 44
    for i in range(1, col + 1):
        temp = PatternFill(patternType='solid', fgColor=Color(indexed=header_colour_code))  # Yellow for header background
        ws[cell_ids[i - 1] + '1'].fill = temp
        ws[cell_ids[i - 1] + '1'].font = temp_font
    # Highlight mismatch rows if any
    if len(missmatch_l2[0]) != 0 or len(missmatch_l2[1]) != 0:
        for i in range(1, col + 1):
            temp = PatternFill(patternType='solid', fgColor=Color(indexed=header_colour_code))  # Yellow background
            ws[cell_ids[i - 1] + str(row + 3)].fill = temp
            ws[cell_ids[i - 1] + str(row + 3)].font = temp_font
    if len(missmatch_l2[0]) != 0 and len(missmatch_l2[1]) != 0:
        for i in range(1, col + 1):
            temp = PatternFill(patternType='solid', fgColor=Color(indexed=header_colour_code))  # Yellow background
            ws[cell_ids[i - 1] + str(row + len(missmatch_l2[0]) + 5)].fill = temp
            ws[cell_ids[i - 1] + str(row + len(missmatch_l2[0]) + 5)].font = temp_font
    # Highlight mismatch score with yellow
    yellow_clr = 5
    for i in range(1, len(mismatch_score_cell_id) + 1):
        col_id = mismatch_score_cell_id[i - 1]
        temp = PatternFill(patternType='solid',
                           fgColor=Color(indexed=yellow_clr))  # Yellow background for mismatch scores
        ws[col_id].fill = temp
    # Save the workbook
    wb.save(file_path)
    # print('Saved file successfully')

    # Save the workbook
    wb.save(file_path)
    print('Saved file successfully')
